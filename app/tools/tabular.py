"""Read CSV or XLSX uploads into the same row-dict shape.

Importers keep their own validation; this module only sniffs the format
and normalises headers so `.csv` and `.xlsx` are interchangeable.
"""

from __future__ import annotations

import csv
import io
from datetime import date, datetime
from typing import Any


class TabularError(ValueError):
    """Unparseable spreadsheet — callers map this to HTTP 400."""


def read_tabular(
    filename: str | None,
    content: bytes,
) -> tuple[list[str], list[tuple[int, dict[str, str]]]]:
    """Return ``(fieldnames, [(row_number, row), ...])``.

    Keys are stripped and lowercased. Values are stripped strings.
    ``row_number`` is 1-based with the header as row 1, matching Excel
    and typical CSV editors.
    """
    if not content:
        raise TabularError("File is empty")

    kind = _sniff_kind(filename, content)
    if kind == "xlsx":
        return _read_xlsx(content)
    return _read_csv(content)


def _sniff_kind(filename: str | None, content: bytes) -> str:
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        return "xlsx"
    if name.endswith(".csv"):
        return "csv"
    # ZIP/XLSX magic
    if content.startswith(b"PK"):
        return "xlsx"
    return "csv"


def _read_csv(content: bytes) -> tuple[list[str], list[tuple[int, dict[str, str]]]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise TabularError(f"CSV must be UTF-8 encoded ({exc.reason})") from exc

    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise TabularError("File has no header row")

    fieldnames = [_norm_header(h) for h in reader.fieldnames]
    if any(not h for h in fieldnames):
        raise TabularError("File has an empty header cell")

    rows: list[tuple[int, dict[str, str]]] = []
    for idx, raw in enumerate(reader, start=2):
        row = {
            _norm_header(k): (v or "").strip() if isinstance(v, str) else _cell_str(v)
            for k, v in raw.items()
            if k is not None
        }
        if not any(row.values()):
            continue
        rows.append((idx, row))
    return fieldnames, rows


def _read_xlsx(content: bytes) -> tuple[list[str], list[tuple[int, dict[str, str]]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise TabularError("Excel support requires openpyxl") from exc

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise TabularError(f"Could not read Excel file: {exc}") from exc

    try:
        ws = wb.active
        if ws is None:
            raise TabularError("Excel file has no active sheet")

        iterator = ws.iter_rows(values_only=True)
        try:
            header_row = next(iterator)
        except StopIteration:
            raise TabularError("File has no header row") from None

        fieldnames = [_norm_header(_cell_str(c)) for c in header_row]
        if not fieldnames or all(not h for h in fieldnames):
            raise TabularError("File has no header row")
        if any(not h for h in fieldnames):
            raise TabularError("File has an empty header cell")

        rows: list[tuple[int, dict[str, str]]] = []
        for excel_row, values in enumerate(iterator, start=2):
            cells = list(values) if values is not None else []
            # Pad / trim to header width
            if len(cells) < len(fieldnames):
                cells.extend([None] * (len(fieldnames) - len(cells)))
            row = {
                fieldnames[i]: _cell_str(cells[i])
                for i in range(len(fieldnames))
            }
            if not any(row.values()):
                continue
            rows.append((excel_row, row))
        return fieldnames, rows
    finally:
        wb.close()


def _norm_header(raw: str | None) -> str:
    return (raw or "").strip().lower()


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
